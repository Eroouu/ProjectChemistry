import openpyxl
import math
import os
import scipy
import numpy as np
import matplotlib.pyplot as plt
import TackensMethof as tm
from numpy import linalg as LA
from mpl_toolkits.mplot3d import Axes3D
from sklearn.decomposition import PCA


def find_file_from(name, x, y):
    wb = openpyxl.load_workbook(filename=name)
    sheet = wb['1']
    time_series, time_array = [], []
    while sheet.cell(row=x, column=y).value is not None:
        time_series.append(float(sheet.cell(row=x, column=y).value))
        time_array.append(float(sheet.cell(row=x, column=y - 1).value))
        x += 1
    return np.array(time_series)


def check_the_probability_to_analyse(matrix: object, n: object, num_n: object, num_e: object, epsilon: object, ind_start: object, ind_end: object) -> object:
    fig = plt.figure(figsize=(15, 10))
    ax = fig.add_subplot(projection="3d")
    nt = np.array([n + i * 5 for i in range(num_n)])
    et = np.array([epsilon * 0.85 ** j for j in range(num_e)])
    rez_arr = np.array([
        [tm.find_spec_funk_value(matrix[ind_start: ind_end], nt[i], et[j]) for j in range(num_e)]
        for i in range(int(num_n))
    ])
    ax.scatter(np.repeat(nt, num_e), np.tile(et, num_n), rez_arr.ravel())
    ax.set_xlabel("n")
    ax.set_ylabel("e")
    plt.title('grapthics for Cne')
    plt.show()


def ex4(matrix, n, num_n, num_e, epsilon, ind_start, ind_end):
    fig = plt.figure(figsize=(12, 7))
    nt = np.array([n + i * 15 for i in range(num_n)])
    et = np.array([epsilon * 0.85 ** j for j in range(num_e)])
    rez_arr = np.array([
        [tm.find_spec_funk_value(matrix[ind_start: ind_end], nt[i], et[j]) for j in range(num_e)]
        for i in range(int(num_n))
    ])
    plt.plot(et, rez_arr[0], 'r', label='n = 20')
    plt.plot(et, rez_arr[1], 'g', label='n = 35')
    plt.plot(et, rez_arr[2], 'b', label='n = 50')
    plt.plot(et, rez_arr[3], 'y', label='n = 65')
    plt.plot(et, rez_arr[4], 'o', label='n = 80')
    plt.legend(bbox_to_anchor=(1.05, 2), loc=2, borderaxespad=0.)
    plt.show()


def try_to_find_main_components(full_series, k, ind_start, ind_end):
    matrix = np.array(full_series[ind_start:ind_end])
    tm.PCA(matrix, k)


def random_series(a, b, length):
    return np.random.rand(length) * (b - a) + a


def logistic_model(x0, mu, numx):
    """
    :param x0:
    :param mu:
    :param numx:
    :return:
    """
    massive = [x0]
    for i in range(numx):
        x1 = mu * x0 * (1 - x0)
        massive.append(x1)
        x0 = x1
    x = [i for i in range(numx+1)]
    plt.plot(x, np.array(massive))
    plt.show()
    return np.array(massive)


def xenon_model(x0, x1, a, b, n):
    massive = [x0, x1]
    print(x0)
    print(x1)
    for i in range(n):
        x2 = 1 - a * x1**2 + b * x0
        massive.append(x2)
        x0 = x1
        x1 = x2
    plt.plot(massive)
    plt.show()
    return np.array(massive)


def try_to_find_main_components_2(full_series, k, ind_start, ind_end):
    matrix = np.array(full_series[ind_start:ind_end])
    tm.PCA_2(matrix, k)


def trying_with_random_value():
    """
    функция проверяет значения Cne|logn + e и тд на рандомном ряде
    :return:
    """
    check_the_probability_to_analyse(random_series(0, 1, 10000), 5, 4, 10, 1, 1000, 3000)


def make_plots_of_components_value(full_series, k, how_many_k_show, ind_start, ind_end):
    series = np.array(full_series[ind_start:ind_end])
    vector_projection = tm.make_array_of_components_value(series, k)
    for i in range(how_many_k_show):
        plt.plot(vector_projection[i], label=f"Компонента = {i}")
        plt.legend()
        plt.show()


def make_3d_cloud_of_points(number_of_points):
    return np.random.random(number_of_points)


def experiment_in_3_dimension(number_of_points):
    x = make_3d_cloud_of_points(number_of_points)
    y = x * 2 + 1 + np.random.random(number_of_points) * 3
    z = 2 * x + 3 * y + np.random.random(number_of_points) * 6
    matrix = np.dstack((x, y, z))
    centered_matrix = np.dstack((x - x.mean(), y - y.mean(), z - z.mean()))
    covmat = np.cov((x - x.mean(), y - y.mean(), z - z.mean()))
    print(covmat)
    eigenvalues, eigenvectors = LA.eig(covmat)
    print(eigenvectors)
    fig = plt.figure(figsize=(4, 4))
    ax = fig.add_subplot(111, projection='3d')
    ax.scatter(x, y, z)
    ax.plot([np.mean(x), np.mean(x) + eigenvectors[0][0]], [np.mean(y), np.mean(y) + eigenvectors[1][0]],
            zs=[np.mean(z), np.mean(z) + eigenvectors[2][0]], color='green')
    ax.plot([np.mean(x), np.mean(x) + eigenvectors[0][1]], [np.mean(y), np.mean(y) + eigenvectors[1][1]],
            zs=[np.mean(z), np.mean(z) + eigenvectors[2][1]], color='orange')
    ax.plot([np.mean(x), np.mean(x) + eigenvectors[0][2]], [np.mean(y), np.mean(y) + eigenvectors[1][2]],
            zs=[np.mean(z), np.mean(z) + eigenvectors[2][2]], color='purple')
    plt.show()
    return 0


def van_der_pol(x0, x1, n, mu):
    matrix_x = []
    matrix_dx = []
    matrix_x.append(x0)
    matrix_dx.append(x1)
    h = 100 / n
    for i in range(n):
        q0 = mu * (1 - x0**2) * x1 - x0
        k0 = x1
        q1 = mu * (1 - (x0+h/2)**2) * (x1 + k0 * h / 2) - (x0+h/2)
        k1 = x1 + q0 * h / 2
        q2 = mu * (1 - (x0+h/2)**2) * (x1 + k1 * h / 2) - (x0+h/2)
        k2 = x1 + q1 * h / 2
        q3 = mu * (1 - (x0+h)**2) * (x1 + k2 * h) - (x0+h)
        k3 = x1 + q2 * h
        x1 = x1 + (q0 + 2 * q1 + 2 * q2 + q3) * h / 6
        x0 = x0 + (k0 + 2 * k1 + 2 * k2 + k3) * h / 6
        matrix_x.append(x0)
        matrix_dx.append(x1)
    x = [i for i in range(n + 1)]
    plt.plot(x, matrix_x)
    plt.show()
    return np.array(matrix_x)

def func_of_circle(x0, x1, r, n ):
    points_x = np.array([(x0 + r * np.cos(i/4)) for i in range(n)])
    points_y =np.array([(x1 + r * np.sin(i/4)) for i in range(n)])
    rez = np.sin(points_x) + points_y
    plt.scatter(points_x, points_y)
    plt.show()
    plt.plot(np.sin(points_x))
    plt.plot(points_y)
    plt.show()
    return rez
if __name__ == '__main__':
    '''
    check_the_probability_to_analyse(find_file_from('Алюминий 2 серия.xlsx', 2, 2), 20, 0.005, 5500, 8500)
    ex4('Алюминий 2 серия.xlsx', 2, 2, 20, 0.0009, 5500, 8500)
    try_to_find_main_components(find_file_from('Алюминий 2 серия.xlsx', 2, 2), 5, 5500, 8500)
    try_to_find_main_components(random_series(10000), 5, 5500, 8500)
    trying_with_random_value()
    ex4(Logistic_Model(0.2,10000), 10, 1, 1, 0.001, 100, 1000)
    check_the_probability_to_analyse(logistic_model(0.2, 4, 10000), 3, 2, 20, 1, 1000, 2000)
    check_the_probability_to_analyse(xenon_model(-0.877, 0.257, 1.8, -0.005, 10000), 5, 4, 20, 1, 1000, 2000)
    check_the_probability_to_analyse(xenon_model(-0.877, 0.257, 1.49, -0.138, 10000), 5, 4, 20, 1, 1000, 2000)
    try_to_find_main_components_2(find_file_from('Алюминий 2 серия.xlsx', 2, 2), 10, 5500, 8500)
    make_plots_of_components_value(find_file_from('Алюминий 2 серия.xlsx', 2, 2), 300, 1, 5500, 8500)
    experiment_in_3_dimension(100)
    check_the_probability_to_analyse(van_der_pol(0.01, 0.01, 1000, 0.5), 10, 6, 30, 1, 1000, 4000)
    '''
    #check_the_probability_to_analyse(van_der_pol(0.01, 0.01, 1000, 0.5), 10, 6, 30, 1, 100, 400)
    #check_the_probability_to_analyse(xenon_model(-0.877, 0.257, 1.8, -0.005, 10000), 5, 4, 20, 0.2, 1000, 2000)
    #try_to_find_main_components_2(van_der_pol(0.01, 0.01, 10000, 0.5), 4, 3000, 7000)
    #try_to_find_main_components_2(find_file_from('Алюминий 2 серия.xlsx', 2, 2), 10, 5500, 8500)
    #experiment_in_3_dimension(100)
    #make_plots_of_components_value(find_file_from('Алюминий 2 серия.xlsx', 2, 2), 300, 1, 5500, 8500)
    make_plots_of_components_value(func_of_circle(0.1, 2, 3, 1000), 40, 2, 0, 1000)